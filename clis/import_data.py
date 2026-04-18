import os
import io
from pathlib import Path
from decimal import Decimal

import django
import openpyxl
from django.db import transaction


os.environ.setdefault("DJANGO_SETTINGS_MODULE", "clis.settings")
django.setup()

from myapp.models import Category, Chemical, ChemicalSub, Item, ItemSub


BASE_DIR = Path(__file__).resolve().parent
LEGACY_DIR = BASE_DIR / "legacy_data"


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def parse_int(value):
    if value is None:
        return 0
    text = str(value).strip().replace(",", "")
    if not text:
        return 0
    try:
        return int(float(text))
    except ValueError:
        return 0


def load_workbook_from_misnamed_xlsx(file_path):
    # Files in legacy_data are xlsx content with .csv extension.
    payload = io.BytesIO(file_path.read_bytes())
    return openpyxl.load_workbook(payload, data_only=True, read_only=True)


def find_header_row(ws, markers):
    for row_idx in range(1, ws.max_row + 1):
        row_values = [normalize_text(cell.value).upper() for cell in ws[row_idx]]
        if all(any(marker in col for col in row_values) for marker in markers):
            return row_idx
    return None


def chemical_total_quantity(row_values, start_col):
    # Legacy sheet stores chemical stock in repeated pairs: [count, amount].
    total = 0
    col = start_col
    while col + 1 < len(row_values):
        count = parse_int(row_values[col])
        amount = parse_int(row_values[col + 1])
        if count > 0 and amount > 0:
            total += count * amount
        col += 2
    return total


@transaction.atomic
def import_chemicals(file_name="chemicals.csv"):
    file_path = LEGACY_DIR / file_name
    if not file_path.exists():
        raise FileNotFoundError(f"Missing file: {file_path}")

    wb = load_workbook_from_misnamed_xlsx(file_path)
    try:
        sheet_name = "STOCK DETAILS OF CHEMICALS"
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

        header_row = find_header_row(ws, ["CHEMICAL NO", "PASSWORD123"])
        if header_row is None:
            raise ValueError("Could not locate chemical header row")

        headers = [normalize_text(c.value).upper() for c in ws[header_row]]
        code_idx = next(i for i, h in enumerate(headers) if "CHEMICAL NO" in h)
        name_idx = next(i for i, h in enumerate(headers) if "PASSWORD123" in h)
        data_start = header_row + 1

        created_chem = 0
        updated_chem = 0
        created_sub = 0
        updated_sub = 0

        for row in ws.iter_rows(min_row=data_start, values_only=True):
            row_values = list(row)
            if len(row_values) <= max(code_idx, name_idx):
                continue

            code = normalize_text(row_values[code_idx])
            name = normalize_text(row_values[name_idx])
            if not code or not name:
                continue

            if "CHEMICAL NO" in code.upper() or "PASSWORD123" in name.upper():
                continue

            category = "organic" if code.upper().startswith("OR") else "inorganic"

            chemical, created = Chemical.objects.get_or_create(
                name=name,
                defaults={
                    "location": "Legacy Stock",
                    "category": category,
                },
            )
            if created:
                created_chem += 1
            else:
                changed = False
                if not chemical.location:
                    chemical.location = "Legacy Stock"
                    changed = True
                if chemical.category not in {"organic", "inorganic"}:
                    chemical.category = category
                    changed = True
                if changed:
                    chemical.save(update_fields=["location", "category"])
                    updated_chem += 1

            total_quantity = chemical_total_quantity(row_values, name_idx + 1)
            if total_quantity <= 0:
                continue

            sub, sub_created = ChemicalSub.objects.update_or_create(
                code=code,
                defaults={
                    "chem_id": chemical,
                    "company": "Legacy Stock",
                    "fund": "Legacy",
                    "quantity": total_quantity,
                    "nos": 1,
                    "exp_date": None,
                },
            )
            sub.total_quantity = sub.nos * sub.quantity
            sub.save(update_fields=["total_quantity"])

            if sub_created:
                created_sub += 1
            else:
                updated_sub += 1

        print("--- Chemicals Import Complete ---")
        print(f"Created chemicals: {created_chem}")
        print(f"Updated chemicals: {updated_chem}")
        print(f"Created chemical stock rows: {created_sub}")
        print(f"Updated chemical stock rows: {updated_sub}")
    finally:
        wb.close()


def glassware_total_units(row_values, start_col):
    # In this sheet, each numeric cell after the descriptor columns represents stock count.
    total = 0
    for value in row_values[start_col:]:
        total += parse_int(value)
    return max(total, 0)


@transaction.atomic
def import_glassware(file_name="glassware.csv"):
    file_path = LEGACY_DIR / file_name
    if not file_path.exists():
        raise FileNotFoundError(f"Missing file: {file_path}")

    wb = load_workbook_from_misnamed_xlsx(file_path)
    try:
        ws = wb["2020-23 DUES LIST"] if "2020-23 DUES LIST" in wb.sheetnames else wb[wb.sheetnames[0]]

        header_row = find_header_row(ws, ["GLASS WARE NO", "PASSWORD123"])
        if header_row is None:
            raise ValueError("Could not locate glassware header row")

        headers = [normalize_text(c.value).upper() for c in ws[header_row]]
        code_idx = next(i for i, h in enumerate(headers) if "GLASS WARE NO" in h)
        name_idx = next(i for i, h in enumerate(headers) if "PASSWORD123" in h)
        data_start = header_row + 1

        glassware_category, _ = Category.objects.get_or_create(name="Glassware")

        created_items = 0
        updated_items = 0
        created_sub_items = 0

        for row in ws.iter_rows(min_row=data_start, values_only=True):
            row_values = list(row)
            if len(row_values) <= max(code_idx, name_idx):
                continue

            base_code = normalize_text(row_values[code_idx])
            name = normalize_text(row_values[name_idx])
            if not base_code or not name:
                continue
            if "GLASS WARE NO" in base_code.upper() or "PASSWORD123" in name.upper():
                continue

            item, created = Item.objects.get_or_create(
                name=name,
                defaults={
                    "specification": "Legacy stock import",
                    "location": "Legacy Store",
                    "category": glassware_category,
                },
            )
            if created:
                created_items += 1
            else:
                changed = False
                if item.category_id != glassware_category.id:
                    item.category = glassware_category
                    changed = True
                if not item.location:
                    item.location = "Legacy Store"
                    changed = True
                if changed:
                    item.save(update_fields=["category", "location"])
                    updated_items += 1

            desired_units = glassware_total_units(row_values, name_idx + 1)
            if desired_units <= 0:
                continue

            existing_units = ItemSub.objects.filter(item_id=item, item_code__startswith=f"{base_code}-").count()
            to_create = max(0, desired_units - existing_units)
            if to_create == 0:
                continue

            bulk_rows = []
            start_no = existing_units + 1
            for i in range(start_no, start_no + to_create):
                bulk_rows.append(
                    ItemSub(
                        item_id=item,
                        item_code=f"{base_code}-{i:04d}",
                        company="Legacy Stock",
                        fund="Legacy",
                        condition="working",
                        price=Decimal("0.00"),
                    )
                )

            ItemSub.objects.bulk_create(bulk_rows, batch_size=500)
            item.update_quantity()
            created_sub_items += to_create

        print("--- Glassware Import Complete ---")
        print(f"Created items: {created_items}")
        print(f"Updated items: {updated_items}")
        print(f"Created item stock rows: {created_sub_items}")
    finally:
        wb.close()


if __name__ == "__main__":
    print("Starting legacy import...")
    import_chemicals()
    import_glassware()
    print("Legacy import completed.")